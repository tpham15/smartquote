import base64
import io
import importlib.util
from openpyxl import Workbook, load_workbook

spec = importlib.util.spec_from_file_location('excel_template', 'api/excel-template.py')
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

wb = Workbook()
ws = wb.active
ws.title = 'Bao gia'
ws['A6'] = 'Khách hàng: KHÁCH HÀNG MẪU CŨ'
ws['E6'] = 'Số báo giá: OLD'
ws['E7'] = 'Ngày: 01/01/2026'
ws['A9'] = 'Địa điểm công trình: Khu vực mẫu'
ws['A10'] = 'Hạng mục: Giải pháp cũ'
headers = ['STT','Khu vực lắp đặt','Tên hàng hoá/ Mô tả','Thông số kỹ thuật/Tính năng cơ bản','Hình ảnh','Mã thiết bị','Xuất xứ','ĐVT','Số lượng','Đơn giá','Thành tiền','Ghi chú']
for idx, header in enumerate(headers, 1):
    ws.cell(13, idx).value = header
ws['C14'] = 'I./ Giải pháp cũ'
ws['K14'] = '=SUM(K15:K17)'
for row in range(15, 18):
    ws.cell(row, 1).value = row - 14
    ws.cell(row, 3).value = f'OLD ITEM {row}'
    ws.cell(row, 6).value = f'OLD-{row}'
    ws.cell(row, 9).value = 1
    ws.cell(row, 10).value = 100
    ws.cell(row, 11).value = 100
ws['A20'] = 'TỔNG HỢP CÁC GIẢI PHÁP NHÀ THÔNG MINH'
ws['A26'] = 'Tổng tiền hàng'
ws['K26'] = 300
ws['A27'] = 'Nhân công thi công'
ws['K27'] = 30
ws['A28'] = 'Tổng giá trị hợp đồng'
ws['K28'] = 330
buf = io.BytesIO()
wb.save(buf)
data_url = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,' + base64.b64encode(buf.getvalue()).decode()

payload = {
    'template': {'fileName': 'demo.xlsx', 'dataUrl': data_url, 'mapping': {
        'sheetName': 'Bao gia',
        'fields': {'customerName': 'A6', 'quoteNumber': 'E6', 'quoteDate': 'E7', 'projectAddress': 'A9', 'projectName': 'A10'},
        'fieldPrefixes': {'customerName': 'Khách hàng: ', 'quoteNumber': 'Số báo giá: ', 'quoteDate': 'Ngày: ', 'projectAddress': 'Địa điểm công trình: ', 'projectName': 'Hạng mục: '},
        'items': {'startRow': 15, 'templateRow': 15, 'clearUntilRow': 19, 'columns': {'no': 'A', 'name': 'C', 'sku': 'F', 'qty': 'I', 'unitPrice': 'J', 'lineTotal': 'K'}},
        'totals': {'subtotal': 'K26', 'labor': 'K27', 'grandTotal': 'K28'},
    }},
    'company': {},
    'customer': {'name': 'Khách Mẫu', 'quoteNumber': 'BG-NEW', 'address': 'Khu vực mới', 'project': 'Công trình mẫu'},
    'calc': {'deviceTotal': 1000, 'laborTotal': 100, 'grand': 1100},
    'rows': [{'no': 1, 'name': 'NEW ITEM', 'sku': 'NEW-1', 'qty': 2, 'unitPrice': 500, 'lineTotal': 1000}],
}

payload["exportMode"] = "template_fidelity_v2"
out = mod._build_workbook(payload)
wb2 = load_workbook(io.BytesIO(out), data_only=False)
ws2 = wb2['Bao gia']
assert ws2['A6'].value == 'Khách hàng: Khách Mẫu'
assert ws2['E6'].value == 'Số báo giá: BG-NEW'
assert ws2['C15'].value == 'NEW ITEM'
assert ws2['C16'].value in (None, '')
assert ws2['C17'].value in (None, '')
assert ws2['K26'].value == 1000
assert ws2['K28'].value == 1100
print('phase 12.1 api workbook smoke: PASS')
