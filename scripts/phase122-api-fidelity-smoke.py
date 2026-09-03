#!/usr/bin/env python3
import base64
import importlib.util
import io
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage

ROOT = Path(__file__).resolve().parents[1]
spec = importlib.util.spec_from_file_location("excel_template", ROOT / "api" / "excel-template.py")
mod = importlib.util.module_from_spec(spec)
sys.modules["excel_template"] = mod
spec.loader.exec_module(mod)


def png_data_url(color=(255, 0, 0, 255)):
    bio = io.BytesIO()
    img = PILImage.new("RGBA", (80, 80), color)
    img.save(bio, format="PNG")
    return "data:image/png;base64," + base64.b64encode(bio.getvalue()).decode()


def make_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bao gia"
    ws.merge_cells("A1:D3")
    ws["A1"] = "LOGO"
    ws["A1"].font = Font(bold=True, size=22, color="1E37B8")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A5"] = "Khách hàng: KHÁCH HÀNG MẪU CŨ"
    ws["E5"] = "Số báo giá: OLD"
    ws["A6"] = "Điện thoại: 090"
    ws["E6"] = "Ngày: 01/01/2026"
    ws["A7"] = "Địa điểm công trình: Khu vực mẫu"
    ws["A8"] = "Hạng mục: Giải pháp nhà thông minh Lumi"
    headers = ["STT", "Khu vực lắp đặt", "Tên hàng hoá/Mô tả", "Thông số kỹ thuật", "Hình ảnh", "Mã thiết bị", "Xuất xứ", "ĐVT", "Số lượng", "Đơn giá", "Thành tiền", "Ghi chú"]
    for idx, title in enumerate(headers, 1):
        c = ws.cell(10, idx)
        c.value = title
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E2F0D9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.row_dimensions[10].height = 30
    ws.merge_cells("A11:J11")
    ws["A11"] = "I./ Giải pháp cũ"
    ws["K11"] = "=SUM(K12:K13)"
    for col in range(1, 13):
        cell = ws.cell(11, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.row_dimensions[11].height = 24
    for r in [12, 13]:
        ws.row_dimensions[r].height = 70
        for col in range(1, 13):
            cell = ws.cell(r, col)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", size=11)
        ws.cell(r, 1).value = r - 11
        ws.cell(r, 3).value = f"Sản phẩm cũ {r}"
        ws.cell(r, 5).value = "https://old.example/image.jpg"
        ws.cell(r, 6).value = f"OLD-{r}"
        ws.cell(r, 7).value = "Lumi"
        ws.cell(r, 8).value = "Cái"
        ws.cell(r, 9).value = 1
        ws.cell(r, 10).value = 1000
        ws.cell(r, 11).value = "=I{}*J{}".format(r, r)
    # stale old image in data region should be removed
    tmp_img = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    PILImage.new("RGB", (40, 40), (0, 0, 255)).save(tmp_img.name)
    ws.add_image(XLImage(tmp_img.name), "E12")
    ws["A15"] = "TỔNG HỢP"
    ws["J16"] = "Tổng tiền hàng"
    ws["K16"] = "=SUM(K11)"
    ws["J17"] = "Nhân công"
    ws["K17"] = "=K16*0.1"
    ws["J18"] = "Tổng giá trị hợp đồng"
    ws["K18"] = "=K16+K17"
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["C"].width = 26
    bio = io.BytesIO()
    wb.save(bio)
    try:
        os.unlink(tmp_img.name)
    except Exception:
        pass
    return "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64," + base64.b64encode(bio.getvalue()).decode()


def main():
    template_url = make_template()
    payload = {
        "template": {
            "fileName": "mau.xlsx",
            "dataUrl": template_url,
            "mapping": {
                "sheetName": "Bao gia",
                "fields": {"customerName": "A5", "customerPhone": "A6", "projectAddress": "A7", "projectName": "A8", "quoteDate": "E6", "quoteNumber": "E5"},
                "fieldPrefixes": {"customerName": "Khách hàng: ", "customerPhone": "Điện thoại: ", "projectAddress": "Địa điểm công trình: ", "projectName": "Hạng mục: ", "quoteDate": "Ngày: ", "quoteNumber": "Số báo giá: "},
                "items": {"startRow": 12, "templateRow": 12, "sectionRow": 11, "clearUntilRow": 13, "columns": {"no": "A", "note": "B", "name": "C", "specs": "D", "image": "E", "sku": "F", "supplier": "G", "unit": "H", "qty": "I", "unitPrice": "J", "lineTotal": "K"}},
                "totals": {"subtotal": "K16", "labor": "K17", "grandTotal": "K18"},
            },
        },
        "company": {},
        "customer": {"name": "Khách Mẫu", "phone": "0900000001", "address": "Khu vực mới", "project": "Công trình mẫu", "quoteNumber": "BG-NEW"},
        "calc": {"deviceTotal": 3000000, "laborTotal": 300000, "grand": 3300000},
        "sections": [
            {"name": "I./ Giải pháp chiếu sáng", "rows": [
                {"no": 1, "note": "Phòng khách", "name": "Công tắc Lumi 1 nút", "specs": "BLE Mesh", "image": png_data_url(), "sku": "LM-1", "supplier": "Lumi", "unit": "Cái", "qty": 2, "unitPrice": 1000000},
                {"no": 2, "note": "Phòng khách", "name": "Công tắc Lumi 2 nút", "specs": "BLE Mesh", "image": "https://example.com/not-used.jpg", "sku": "LM-2", "supplier": "Lumi", "unit": "Cái", "qty": 1, "unitPrice": 1000000},
            ]},
            {"name": "II./ Hệ thống camera", "rows": [
                {"no": 3, "note": "Ngoài cổng", "name": "Camera Hikvision", "specs": "AI", "sku": "DS-2", "supplier": "Hikvision", "unit": "Cái", "qty": 1, "unitPrice": 1000000},
            ]},
        ],
        "exportMode": "template_fidelity_v2",
    }
    out = mod._build_workbook(payload)
    wb = load_workbook(io.BytesIO(out), data_only=False)
    ws = wb["Bao gia"]
    assert ws["A1"].value == "LOGO", "header/logo text not preserved"
    assert "A1:D3" in [str(r) for r in ws.merged_cells.ranges], "header merge not preserved"
    assert ws["A5"].value == "Khách hàng: Khách Mẫu", "field prefix not preserved"
    assert ws["A11"].value == "I./ Giải pháp chiếu sáng", "section row not written"
    assert "A11:J11" in [str(r) for r in ws.merged_cells.ranges], "section merge not recreated"
    assert ws["C12"].value == "Công tắc Lumi 1 nút", "first item not written"
    assert ws["E12"].value in (None, ""), "image URL/text should not be written into image cell"
    assert len(ws._images) >= 1, "data URL image should be inserted"
    assert str(ws["K12"].value).startswith("=I12*J12"), "line total formula missing"
    assert ws["A14"].value == "II./ Hệ thống camera", "second section should be inserted after first two items"
    assert ws["K21"].value and str(ws["K21"].value).startswith("="), "shifted grand total formula missing"
    assert ws.row_dimensions[12].height == 70, "item row height not copied"
    print("Phase 12.2 Excel fidelity smoke: PASS")


if __name__ == "__main__":
    main()
