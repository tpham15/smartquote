"""
Phase 12.4 — Lossless Excel Template Engine
POST /api/excel-template

Primary export path (`lossless_xml_v3`) preserves the original XLSX package and patches
only mapped OOXML parts. Phase 12.2's openpyxl path remains as an explicit legacy mode
for backward compatibility and regression coverage; the current frontend does not
silently fall back to it.
"""
from http.server import BaseHTTPRequestHandler
import base64
import copy as pycopy
import io
import ipaddress
import json
import os
import re
import socket
import sys
import time
from datetime import date
from urllib.parse import urlparse

sys.path.insert(0, os.path.dirname(__file__))
from auth_guard import (
    require_api_access, assert_within_quota, assert_rate_limit, record_usage,
    cors_origin, origin_allowed, security_headers,
)

from openpyxl import load_workbook
from xlsx_lossless import analyze_template, build_lossless_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries
try:
    from PIL import Image as PILImage
except Exception:  # pragma: no cover - Pillow is in requirements, fallback keeps API importable.
    PILImage = None
try:
    import requests
except Exception:  # pragma: no cover
    requests = None


MAX_TEMPLATE_BYTES = int(os.getenv("SMARTQUOTE_MAX_EXCEL_TEMPLATE_BYTES", "8000000"))
MAX_IMAGE_BYTES = int(os.getenv("SMARTQUOTE_MAX_EXCEL_IMAGE_BYTES", "1500000"))
MAX_REMOTE_IMAGES = int(os.getenv("SMARTQUOTE_MAX_EXCEL_REMOTE_IMAGES", "40"))


def _json_response(handler, status, payload):
    handler.send_response(status)
    _cors(handler)
    handler.send_header("Content-Type", "application/json")
    handler.end_headers()
    handler.wfile.write(json.dumps(payload).encode("utf-8"))


def _cors(handler):
    origin = cors_origin(handler.headers)
    if origin:
        handler.send_header("Access-Control-Allow-Origin", origin)
    handler.send_header("Vary", "Origin")
    handler.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
    handler.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization, X-SmartQuote-Dealer-Id, X-SmartQuote-Event, X-SmartQuote-Units, X-Request-Id")
    handler.send_header("Access-Control-Max-Age", "86400")
    handler.send_header("Access-Control-Expose-Headers", "X-SmartQuote-Excel-Engine, X-SmartQuote-Excel-Manifest, X-SmartQuote-Excel-Source-SHA256")
    security_headers(handler)


def _decode_data_url(data_url: str) -> bytes:
    if not data_url or not isinstance(data_url, str):
        raise ValueError("Mẫu Excel chưa có dữ liệu file.")
    if "," in data_url:
        data_url = data_url.split(",", 1)[1]
    return base64.b64decode(data_url)


def _safe_sheet(wb, name: str):
    if name and name in wb.sheetnames:
        return wb[name]
    return wb[wb.sheetnames[0]]


def _cell_ref(value):
    value = str(value or "").strip().upper()
    if not re.match(r"^[A-Z]{1,3}[0-9]{1,7}$", value):
        return None
    return value


def _cell_row(ref):
    m = re.search(r"(\d+)$", str(ref or ""))
    return int(m.group(1)) if m else None


def _col_idx(value):
    value = str(value or "").strip().upper()
    if not re.match(r"^[A-Z]{1,3}$", value):
        return None
    return column_index_from_string(value)


def _set(ws, ref, value):
    ref = _cell_ref(ref)
    if not ref:
        return
    ws[ref] = value


def _shift_cell_ref(ref, first_shifted_row: int, delta: int):
    ref = _cell_ref(ref)
    if not ref or not delta:
        return ref
    m = re.match(r"^([A-Z]{1,3})([0-9]{1,7})$", ref)
    if not m:
        return ref
    col, row_s = m.group(1), m.group(2)
    row = int(row_s)
    if row >= first_shifted_row:
        return f"{col}{max(1, row + delta)}"
    return ref


def _set_field(ws, ref, value, prefix=""):
    ref = _cell_ref(ref)
    if not ref:
        return
    if prefix:
        ws[ref] = f"{prefix}{value or ''}"
    else:
        ws[ref] = value


def _clear_row_values(ws, row: int, max_col: int = None):
    max_col = max_col or ws.max_column
    for col in range(1, max_col + 1):
        try:
            _writable_cell(ws, row, col).value = None
        except Exception:
            pass


def _copy_cell_style(src, dst):
    if src.has_style:
        dst.font = pycopy.copy(src.font)
        dst.fill = pycopy.copy(src.fill)
        dst.border = pycopy.copy(src.border)
        dst.alignment = pycopy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = pycopy.copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = pycopy.copy(src.hyperlink)
    if src.comment:
        dst.comment = pycopy.copy(src.comment)


def _capture_row_template(ws, row: int, max_col: int = None):
    max_col = max_col or ws.max_column
    cells = []
    for col in range(1, max_col + 1):
        src = ws.cell(row, col)
        cells.append({
            "col": col,
            "font": pycopy.copy(src.font),
            "fill": pycopy.copy(src.fill),
            "border": pycopy.copy(src.border),
            "alignment": pycopy.copy(src.alignment),
            "number_format": src.number_format,
            "protection": pycopy.copy(src.protection),
            "style_id": src.style_id,
        })
    merges = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col_m, max_row = range_boundaries(str(mr))
        if min_row <= row <= max_row:
            merges.append({
                "row_offset_start": min_row - row,
                "row_offset_end": max_row - row,
                "min_col": min_col,
                "max_col": max_col_m,
            })
    return {
        "row": row,
        "height": ws.row_dimensions[row].height,
        "hidden": ws.row_dimensions[row].hidden,
        "cells": cells,
        "merges": merges,
    }


def _apply_row_template(ws, tpl, target_row: int, clear_values=True):
    if not tpl or target_row <= 0:
        return
    ws.row_dimensions[target_row].height = tpl.get("height")
    ws.row_dimensions[target_row].hidden = tpl.get("hidden", False)
    for cell_tpl in tpl.get("cells", []):
        col = cell_tpl["col"]
        dst = ws.cell(target_row, col)
        dst.font = pycopy.copy(cell_tpl["font"])
        dst.fill = pycopy.copy(cell_tpl["fill"])
        dst.border = pycopy.copy(cell_tpl["border"])
        dst.alignment = pycopy.copy(cell_tpl["alignment"])
        dst.number_format = cell_tpl["number_format"]
        dst.protection = pycopy.copy(cell_tpl["protection"])
        if clear_values:
            dst.value = None


def _recreate_row_merges(ws, tpl, target_row: int, max_existing_row: int = None, horizontal_only: bool = False):
    if not tpl:
        return
    for m in tpl.get("merges", []):
        if horizontal_only and m["row_offset_start"] != m["row_offset_end"]:
            # Product rows in real templates sometimes merge a spec cell vertically across multiple old items.
            # Recreating that merge for generated items would make later rows non-writable.
            continue
        min_row = target_row + m["row_offset_start"]
        max_row = target_row + m["row_offset_end"]
        if min_row <= 0 or max_row <= 0:
            continue
        if max_existing_row and (min_row > max_existing_row or max_row > max_existing_row):
            continue
        ref = f"{get_column_letter(m['min_col'])}{min_row}:{get_column_letter(m['max_col'])}{max_row}"
        try:
            ws.merge_cells(ref)
        except ValueError:
            # Already merged, safe to ignore.
            pass


def _unmerge_intersecting_rows(ws, first_row: int, last_row: int):
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        if max_row >= first_row and min_row <= last_row:
            try:
                ws.unmerge_cells(str(mr))
            except Exception:
                pass


def _anchor_row(img):
    try:
        # openpyxl stores anchor._from.row as zero-based
        return int(img.anchor._from.row) + 1
    except Exception:
        return None


def _remove_images_in_rows(ws, first_row: int, last_row: int):
    if not hasattr(ws, "_images"):
        return 0
    kept = []
    removed = 0
    for img in list(ws._images):
        row = _anchor_row(img)
        if row is not None and first_row <= row <= last_row:
            removed += 1
            continue
        kept.append(img)
    ws._images = kept
    return removed


def _safe_remote_image_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
        if parsed.scheme not in ("http", "https") or not parsed.hostname:
            return False
        host = parsed.hostname.lower()
        if host in ("localhost",) or host.endswith(".local"):
            return False
        # Resolve and block private/link-local/reserved networks. If DNS fails, skip.
        for info in socket.getaddrinfo(host, None):
            ip = ipaddress.ip_address(info[4][0])
            if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_multicast or ip.is_reserved:
                return False
        return True
    except Exception:
        return False


def _image_bytes_from_source(src: str, remote_budget: dict) -> bytes | None:
    src = str(src or "").strip()
    if not src:
        return None
    if src.startswith("data:image/"):
        try:
            return _decode_data_url(src)[:MAX_IMAGE_BYTES + 1]
        except Exception:
            return None
    if src.startswith("http://") or src.startswith("https://"):
        if remote_budget.get("count", 0) >= MAX_REMOTE_IMAGES:
            return None
        if not requests or not _safe_remote_image_url(src):
            return None
        try:
            remote_budget["count"] = remote_budget.get("count", 0) + 1
            r = requests.get(src, timeout=5, stream=True, headers={"User-Agent": "SmartQuote/1.0"})
            r.raise_for_status()
            ctype = r.headers.get("Content-Type", "")
            if "image" not in ctype.lower():
                return None
            chunks = []
            total = 0
            for chunk in r.iter_content(8192):
                total += len(chunk)
                if total > MAX_IMAGE_BYTES:
                    return None
                chunks.append(chunk)
            return b"".join(chunks)
        except Exception:
            return None
    return None



def _writable_cell(ws, row: int, col: int):
    # If the target is inside a merged range, write to that range's top-left cell.
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(min_row, min_col)
    return ws.cell(row, col)

def _cell_box_px(ws, row: int, col: int):
    letter = get_column_letter(col)
    width = ws.column_dimensions[letter].width or 10
    # Excel column width approximation in pixels.
    px_w = max(30, int(width * 7 + 5))
    height_pt = ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 36
    px_h = max(24, int(height_pt * 1.333))
    return px_w, px_h


def _insert_image(ws, row: int, col: int, src: str, remote_budget: dict):
    data = _image_bytes_from_source(src, remote_budget)
    if not data or len(data) > MAX_IMAGE_BYTES:
        return False
    if not PILImage:
        return False
    try:
        raw = io.BytesIO(data)
        pil = PILImage.open(raw)
        pil.verify()
        raw.seek(0)
        pil = PILImage.open(raw)
        # Normalize to PNG/JPEG openpyxl-compatible stream.
        out = io.BytesIO()
        if pil.mode not in ("RGB", "RGBA"):
            pil = pil.convert("RGBA")
        fmt = "PNG" if pil.mode == "RGBA" else "JPEG"
        pil.save(out, format=fmt)
        out.seek(0)
        img = XLImage(out)
        box_w, box_h = _cell_box_px(ws, row, col)
        max_w, max_h = max(20, box_w - 8), max(20, box_h - 8)
        scale = min(max_w / max(1, img.width), max_h / max(1, img.height), 1.0)
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, f"{get_column_letter(col)}{row}")
        return True
    except Exception:
        return False


def _flatten_sections(data: dict):
    sections = []
    if isinstance(data.get("sections"), list) and data.get("sections"):
        for sec in data.get("sections"):
            rows = [r for r in (sec.get("rows") or []) if isinstance(r, dict)]
            if rows:
                sections.append({"name": sec.get("name") or rows[0].get("room") or "Hạng mục", "rows": rows})
    if sections:
        return sections
    rows = [r for r in (data.get("rows") or []) if isinstance(r, dict)]
    if not rows:
        return []
    grouped = []
    by_name = {}
    for row in rows:
        name = row.get("room") or row.get("note") or "Hạng mục"
        if name not in by_name:
            by_name[name] = {"name": name, "rows": []}
            grouped.append(by_name[name])
        by_name[name]["rows"].append(row)
    return grouped


def _write_row_values(ws, excel_row: int, row: dict, columns: dict, image_budget: dict):
    for key, col in (columns or {}).items():
        col_i = _col_idx(col)
        if not col_i:
            continue
        cell = _writable_cell(ws, excel_row, col_i)
        if key == "image":
            # Never write long image URLs into quote templates. Insert real images or leave blank.
            cell.value = None
            _insert_image(ws, excel_row, col_i, row.get("image") or row.get("imageUrl") or "", image_budget)
            continue
        if key == "lineTotal":
            qty_col = _col_idx(columns.get("qty"))
            unit_col = _col_idx(columns.get("unitPrice"))
            if qty_col and unit_col:
                cell.value = f"={get_column_letter(qty_col)}{excel_row}*{get_column_letter(unit_col)}{excel_row}"
            else:
                cell.value = float(row.get("lineTotal") or 0)
            continue
        value = row.get(key, "")
        if key in ("qty", "unitPrice", "lineTotal"):
            try:
                value = float(value or 0)
            except Exception:
                value = 0
        cell.value = value


def _build_workbook_v2(data: dict) -> bytes:
    template = data.get("template") or {}
    mapping = template.get("mapping") or {}
    fields = mapping.get("fields") or {}
    field_prefixes = mapping.get("fieldPrefixes") or {}
    items = mapping.get("items") or {}
    columns = items.get("columns") or {}
    totals = mapping.get("totals") or {}
    company = data.get("company") or {}
    customer = data.get("customer") or {}
    calc = data.get("calc") or {}
    sections = _flatten_sections(data)

    if not str(template.get("fileName") or "").lower().endswith(".xlsx") and not str(template.get("dataUrl") or "").startswith("data:"):
        raise ValueError("Phase 12 chỉ hỗ trợ mẫu Excel .xlsx.")

    raw = _decode_data_url(template.get("dataUrl") or "")
    if len(raw) > MAX_TEMPLATE_BYTES:
        raise ValueError("Mẫu Excel quá lớn. Hãy dùng file .xlsx nhỏ hơn.")

    wb = load_workbook(io.BytesIO(raw))
    ws = _safe_sheet(wb, mapping.get("sheetName") or "")

    # Force Excel to recalculate formulas after opening.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    today = date.today().strftime("%d/%m/%Y")
    def write_field(key, value):
        _set_field(ws, fields.get(key), value, field_prefixes.get(key) or "")

    write_field("customerName", customer.get("name") or "")
    write_field("customerPhone", customer.get("phone") or "")
    write_field("projectAddress", customer.get("address") or "")
    write_field("projectName", customer.get("project") or customer.get("category") or "")
    write_field("quoteDate", today)
    write_field("quoteNumber", customer.get("quoteNumber") or company.get("quoteNumber") or "")
    write_field("companyName", company.get("name") or "")
    write_field("salesPerson", company.get("salesPerson") or "")
    write_field("salesPhone", company.get("salesPhone") or company.get("phone") or "")

    start_row = max(1, int(items.get("startRow") or 15))
    template_row = max(1, int(items.get("templateRow") or start_row))
    section_row = int(items.get("sectionRow") or 0) or None
    clear_until_row = max(start_row, int(items.get("clearUntilRow") or (start_row + sum(len(s["rows"]) for s in sections) - 1)))
    data_first_row = min(section_row or start_row, start_row)
    data_last_row = max(clear_until_row, template_row, section_row or 0)
    max_col = max(ws.max_column, max([_col_idx(c) or 0 for c in columns.values()] + [12]))

    item_tpl = _capture_row_template(ws, template_row, max_col=max_col)
    section_tpl = _capture_row_template(ws, section_row, max_col=max_col) if section_row else None

    # Remove old data region entirely so product images, stale values, merged section rows and blank rows do not remain.
    old_count = max(1, data_last_row - data_first_row + 1)
    generated_count = 0
    for section in sections:
        if section_tpl:
            generated_count += 1
        generated_count += len(section["rows"])
    generated_count = max(1, generated_count)

    _remove_images_in_rows(ws, data_first_row, data_last_row)
    structural_edit = bool(section_row) or generated_count > old_count
    if structural_edit:
        _unmerge_intersecting_rows(ws, data_first_row, data_last_row)
        ws.delete_rows(data_first_row, old_count)
        ws.insert_rows(data_first_row, generated_count)
        row_delta = generated_count - old_count
        first_shifted_original_row = data_last_row + 1
    else:
        # Legacy/flat templates: keep summary/footer absolute positions to avoid surprising layout jumps.
        # Clear stale rows in-place and reuse the existing capacity.
        row_delta = 0
        first_shifted_original_row = data_last_row + 1
        for r in range(data_first_row, data_last_row + 1):
            _clear_row_values(ws, r, max_col=max_col)
            _apply_row_template(ws, item_tpl, r, clear_values=True)

    current = data_first_row
    line_total_cells = []
    section_total_cells = []
    image_budget = {"count": 0}

    for section in sections:
        sec_total_refs = []
        if section_tpl:
            _apply_row_template(ws, section_tpl, current, clear_values=True)
            _recreate_row_merges(ws, section_tpl, current)
            name_col = _col_idx(columns.get("name")) or 1
            total_col = _col_idx(columns.get("lineTotal"))
            _writable_cell(ws, current, name_col).value = section.get("name") or "Hạng mục"
            section_total_cell = None
            if total_col:
                section_total_cell = f"{get_column_letter(total_col)}{current}"
                ws.cell(current, total_col).value = 0
                section_total_cells.append(section_total_cell)
            current += 1
        else:
            section_total_cell = None

        first_item_row = current
        for row in section["rows"]:
            _apply_row_template(ws, item_tpl, current, clear_values=True)
            _recreate_row_merges(ws, item_tpl, current, horizontal_only=True)
            _write_row_values(ws, current, row, columns, image_budget)
            total_col = _col_idx(columns.get("lineTotal"))
            if total_col:
                ref = f"{get_column_letter(total_col)}{current}"
                line_total_cells.append(ref)
                sec_total_refs.append(ref)
            current += 1
        if section_tpl and section_total_cell:
            if sec_total_refs:
                ws[section_total_cell] = f"=SUM({','.join(sec_total_refs)})"
            else:
                ws[section_total_cell] = 0

    # If no rows are present, keep one styled blank item row so the template still opens cleanly.
    if not sections:
        _apply_row_template(ws, item_tpl, current, clear_values=True)
        _recreate_row_merges(ws, item_tpl, current)

    subtotal_ref = _shift_cell_ref(totals.get("subtotal"), first_shifted_original_row, row_delta)
    labor_ref = _shift_cell_ref(totals.get("labor"), first_shifted_original_row, row_delta)
    vat_ref = _shift_cell_ref(totals.get("vat"), first_shifted_original_row, row_delta)
    grand_ref = _shift_cell_ref(totals.get("grandTotal"), first_shifted_original_row, row_delta)

    # Rebuild the optional "Tổng hợp các giải pháp" summary rows that live between
    # the data table and the subtotal row in many dealer templates. Leaving old rows
    # here would show stale section names/formulas from the uploaded sample quote.
    subtotal_row = _cell_row(subtotal_ref) if subtotal_ref else None
    total_col_i = _col_idx(columns.get("lineTotal"))
    if structural_edit and section_row and subtotal_row and current < subtotal_row and total_col_i:
        summary_first_row = current + 1  # row after the summary title
        summary_last_row = subtotal_row - 1
        available = max(0, summary_last_row - summary_first_row + 1)
        if available < len(section_total_cells):
            # Add missing summary rows immediately above subtotal. Existing footer shifts down.
            add = len(section_total_cells) - available
            ws.insert_rows(subtotal_row, add)
            row_delta += add
            subtotal_ref = _shift_cell_ref(subtotal_ref, subtotal_row, add)
            labor_ref = _shift_cell_ref(labor_ref, subtotal_row, add)
            vat_ref = _shift_cell_ref(vat_ref, subtotal_row, add)
            grand_ref = _shift_cell_ref(grand_ref, subtotal_row, add)
            subtotal_row += add
            summary_last_row += add
        for idx, sec in enumerate(sections):
            rr = summary_first_row + idx
            if rr >= subtotal_row:
                break
            _clear_row_values(ws, rr, max_col=max_col)
            _writable_cell(ws, rr, 1).value = sec.get("name") or f"Hạng mục {idx + 1}"
            ws.cell(rr, total_col_i).value = f"={section_total_cells[idx]}"
        for rr in range(summary_first_row + len(section_total_cells), summary_last_row + 1):
            _clear_row_values(ws, rr, max_col=max_col)

    if subtotal_ref:
        if not structural_edit and not section_row:
            ws[subtotal_ref] = float(calc.get("deviceTotal") or 0)
        elif section_total_cells:
            ws[subtotal_ref] = f"=SUM({','.join(section_total_cells)})"
        elif line_total_cells:
            ws[subtotal_ref] = f"=SUM({','.join(line_total_cells)})"
        else:
            ws[subtotal_ref] = float(calc.get("deviceTotal") or 0)
    if labor_ref:
        device_total = float(calc.get("deviceTotal") or 0)
        labor_total = float(calc.get("laborTotal") or 0)
        if not structural_edit and not section_row:
            ws[labor_ref] = labor_total
        elif subtotal_ref and device_total > 0 and labor_total > 0:
            ratio = labor_total / device_total
            ws[labor_ref] = f"={subtotal_ref}*{ratio:.6f}"
        else:
            ws[labor_ref] = labor_total
    if vat_ref:
        ws[vat_ref] = 0
    if grand_ref:
        if not structural_edit and not section_row:
            ws[grand_ref] = float(calc.get("grand") or 0)
        else:
            parts = [ref for ref in (subtotal_ref, labor_ref, vat_ref) if ref]
            ws[grand_ref] = f"=SUM({','.join(parts)})" if parts else float(calc.get("grand") or 0)

    # Preserve print area where possible. If the old print area ended below the edited region,
    # openpyxl may not shift it perfectly, so extend it to the current used range as a safe default.
    try:
        if ws.print_area:
            ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_workbook(data: dict) -> bytes:
    """Phase 12.4.2: lossless is the default. Legacy v2 must be requested explicitly."""
    mode = str(data.get("exportMode") or "lossless_xml_v3")
    if mode == "template_fidelity_v2":
        return _build_workbook_v2(data)
    template = data.get("template") or {}
    raw = _decode_data_url(template.get("dataUrl") or "")
    if len(raw) > MAX_TEMPLATE_BYTES:
        raise ValueError("Mẫu Excel quá lớn. Hãy dùng file .xlsx nhỏ hơn.")
    expected_checksum = str(template.get("sourceChecksum") or "").strip().lower()
    if expected_checksum:
        import hashlib
        actual = hashlib.sha256(raw).hexdigest()
        if actual != expected_checksum:
            raise ValueError("File mẫu đã thay đổi so với manifest đã lưu. Hãy phân tích lại mẫu Excel trước khi xuất.")
    image_budget = {"count": 0}
    output, report = build_lossless_workbook(
        raw,
        data,
        mapping_hint=template.get("mapping") or template.get("manifest") or {},
        image_loader=lambda src: _image_bytes_from_source(src, image_budget),
    )
    return output


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        _cors(self)
        self.end_headers()

    def do_POST(self):
        started = time.time()
        if not origin_allowed(self.headers):
            _json_response(self, 403, {"error": "Origin không được phép gọi SmartQuote API."})
            return

        auth, auth_error = require_api_access(self)
        if auth_error:
            status, payload = auth_error
            _json_response(self, status, payload)
            return

        try:
            length = int(self.headers.get("Content-Length", 0))
            max_body = int(os.getenv("SMARTQUOTE_MAX_EXCEL_TEMPLATE_BODY_BYTES", "11000000"))
            if length > max_body:
                _json_response(self, 413, {"error": "File mẫu hoặc báo giá quá lớn. Hãy dùng mẫu Excel nhỏ hơn."})
                return
            data = json.loads(self.rfile.read(length))

            # Phase 12.4: analyze the immutable source workbook without consuming export quota.
            if str(data.get("action") or "").lower() == "analyze":
                template = data.get("template") or {}
                raw = _decode_data_url(template.get("dataUrl") or "")
                if len(raw) > MAX_TEMPLATE_BYTES:
                    _json_response(self, 413, {"error": "Mẫu Excel quá lớn. Hãy dùng file .xlsx nhỏ hơn."})
                    return
                analyzed = analyze_template(raw, (template.get("mapping") or {}).get("sheetName") or "")
                _json_response(self, 200, analyzed)
                return

            rate_error = assert_rate_limit(auth, self, "excel_export", 1)
            if rate_error:
                status, payload = rate_error
                _json_response(self, status, payload)
                return

            quota_error = assert_within_quota(auth, "excel_export", 1)
            if quota_error:
                status, payload = quota_error
                _json_response(self, status, payload)
                return

            output = _build_workbook(data)
            engine_mode = str(data.get("exportMode") or "lossless_xml_v3")
            customer = data.get("customer", {})
            raw_name = str(customer.get("name") or "BaoGia")[:80]
            safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", raw_name).strip("_") or "BaoGia"
            filename = f"BaoGia_{safe_name}_theo_mau.xlsx"

            self.send_response(200)
            _cors(self)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("X-SmartQuote-Excel-Engine", engine_mode)
            if engine_mode == "lossless_xml_v3":
                self.send_header("X-SmartQuote-Excel-Manifest", "3")
                template = data.get("template") or {}
                if template.get("sourceChecksum"):
                    self.send_header("X-SmartQuote-Excel-Source-SHA256", str(template.get("sourceChecksum")))
            self.send_header("X-Content-Type-Options", "nosniff")
            self.end_headers()
            self.wfile.write(output)
            record_usage(auth, "excel_export", 1, {"customer": raw_name, "filename": filename, "template": (data.get("template") or {}).get("name") or "", "mode": data.get("exportMode") or "lossless_xml_v3", "durationMs": int((time.time() - started) * 1000)})
        except Exception as e:
            _json_response(self, 500, {"error": str(e)})
