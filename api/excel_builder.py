import json, sys, io, requests
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from datetime import date

# ── Màu brand xanh lá ──────────────────────────────────────────────────────
BRAND     = "FF1A7A4A"
BRAND_D   = "FF155E3A"
BRAND_BG  = "FFD1FAE5"
HDR_BG    = "FF0F3D26"
SEC_BG    = "FFE8F5E9"
SEC_FG    = "FF1A7A4A"
WHITE     = "FFFFFFFF"
GRAY_BG   = "FFF8FAFC"
GRAY_LINE = "FFE2E8F0"
TEXT      = "FF0F172A"
TEXT2     = "FF475569"

def side(color=GRAY_LINE, style="thin"):
    return Side(border_style=style, color=color)

def border(color=GRAY_LINE):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=TEXT, name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def fetch_img(url, size=(56,56)):
    """Tải ảnh từ URL → openpyxl Image object. Trả None nếu lỗi."""
    if not url or "encrypted-tbn" in url:
        return None
    try:
        r = requests.get(url, timeout=5,
                         headers={"User-Agent":"Mozilla/5.0",
                                  "Referer":"https://google.com/"})
        r.raise_for_status()
        img = PILImage.open(io.BytesIO(r.content)).convert("RGB")
        img.thumbnail(size, PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, "PNG")
        buf.seek(0)
        xl = XLImage(buf)
        xl.width, xl.height = size
        return xl
    except:
        return None

def build_excel(data: dict) -> bytes:
    company  = data.get("company", {})
    customer = data.get("customer", {})
    rooms    = data.get("rooms", [])
    calc     = data.get("calc", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo giá"

    today = date.today().strftime("%d/%m/%Y")
    vnd = lambda n: f"{int(n or 0):,}".replace(",",".")

    # ── Cột widths ───────────────────────────────────────────────────────────
    col_w = [5, 22, 36, 32, 10, 14, 8, 6, 14, 16]
    #        STT KhuVuc Tên  Spec  Ảnh  Mã  NCC ĐVT SL Đơn Thành
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Cột ảnh rộng hơn để chứa thumbnail 56px
    ws.column_dimensions["E"].width = 9

    # Helper merge & write
    def mr(r, c1, c2, val, fnt=None, fil=None, aln=None, brd=None):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(r, c1, val)
        if fnt: cell.font = fnt
        if fil: cell.fill = fil
        if aln: cell.alignment = aln
        if brd:
            for col in range(c1, c2+1):
                ws.cell(r, col).border = brd
        return cell

    def row_h(r, h):
        ws.row_dimensions[r].height = h

    R = 1  # current row counter

    # ══════════════════════════════════════════════════════════════════════════
    # HEADER CÔNG TY
    # ══════════════════════════════════════════════════════════════════════════
    row_h(R, 18)
    mr(R,1,5, company.get("name",""), font(True,13,WHITE), fill(HDR_BG),
       align("center"), border(HDR_BG))
    mr(R,6,10, f"Số báo giá: {customer.get('quoteNumber','')}", font(False,10,WHITE),
       fill(HDR_BG), align("right"), border(HDR_BG))
    R+=1

    row_h(R,14)
    mr(R,1,5, f"Showroom & VPGD: {company.get('address','')}",
       font(False,9,WHITE), fill(BRAND_D), align("left"), border(BRAND_D))
    mr(R,6,10, f"Ngày: {today}", font(False,9,WHITE),
       fill(BRAND_D), align("right"), border(BRAND_D))
    R+=1

    row_h(R,13)
    mr(R,1,5, f"MST: {company.get('taxCode','')}  |  ĐT: {company.get('phone','')}",
       font(False,9,WHITE), fill(BRAND_D), align("left"), border(BRAND_D))
    mr(R,6,10, f"Người báo giá: {company.get('salesPerson','')}",
       font(False,9,WHITE), fill(BRAND_D), align("right"), border(BRAND_D))
    R+=1

    row_h(R,12)
    website = company.get("website","")
    mr(R,1,10, website, font(False,8,"FF86EFAC"), fill(BRAND_D),
       align("center"), border(BRAND_D))
    R+=1

    # Khoảng trống
    R+=1

    # ── Thông tin khách ───────────────────────────────────────────────────────
    info = [
        (f"Khách hàng: {customer.get('name','')}", f"Điện thoại NB: {company.get('salesPhone','')}"),
        (f"Điện thoại: {customer.get('phone','')}", ""),
        (f"Email: ", ""),
        (f"Địa điểm: {customer.get('address','')}", ""),
        (f"Hạng mục: {customer.get('category','Giải pháp nhà thông minh')}", ""),
    ]
    for left, right in info:
        row_h(R,14)
        mr(R,1,5, left, font(False,10), fill(GRAY_BG), align("left"),border())
        mr(R,6,10, right, font(False,10), fill(GRAY_BG), align("right"),border())
        R+=1
    R+=1

    # ── Tiêu đề báo giá ───────────────────────────────────────────────────────
    row_h(R,18)
    mr(R,1,10, "BẢNG BÁO GIÁ TỔNG HỢP", font(True,14,WHITE),
       fill(BRAND), align("center"), border(BRAND))
    R+=1
    row_h(R,28)
    mr(R,1,10,
       "Công ty cổ phần Nguyên Đà group trân trọng cảm ơn sự quan tâm của Quý Khách hàng "
       "tới sản phẩm của công ty chúng tôi. Xin được gửi tới Quý Khách hàng Bảng báo giá với những chi tiết như sau:",
       font(False,9,TEXT2), fill(GRAY_BG), align("left",wrap=True), border())
    R+=1

    # ── Header bảng ───────────────────────────────────────────────────────────
    headers = ["STT","Khu vực lắp đặt","Tên hàng hoá / Mô tả",
               "Thông số kỹ thuật","Hình ảnh","Mã thiết bị","Xuất xứ","ĐVT","SL","Đơn giá","Thành tiền"]
    row_h(R,20)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(R, c, h)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = fill(BRAND)
        cell.alignment = align("center")
        cell.border = border(BRAND)
    # Cột 11 = Thành tiền
    ws.column_dimensions[get_column_letter(11)].width = 16
    R+=1

    # ── Rows thiết bị ─────────────────────────────────────────────────────────
    section_rows = []  # (tên, excel_row) cho bảng tổng hợp

    for room in rooms:
        lines = [l for l in room.get("lines",[]) if l.get("productId")]
        if not lines:
            continue

        # Dòng tiêu đề giải pháp
        sec_start = R
        row_h(R,16)
        sec_name = room.get("name","").replace("\n"," ")
        for c in range(1,12):
            cell = ws.cell(R,c)
            cell.fill = fill(SEC_BG)
            cell.border = border(SEC_FG)
        ws.merge_cells(start_row=R,start_column=1,end_row=R,end_column=10)
        ws.cell(R,1, sec_name).font = Font(name="Arial",bold=True,size=10,color=SEC_FG)
        ws.cell(R,1).alignment = align("left")
        # Tổng giải pháp (công thức sẽ điền sau)
        sec_total_row = R
        section_rows.append((sec_name, R, []))
        R+=1

        stt = 0
        for l in lines:
            p     = l.get("product", {})
            stt  += 1
            qty   = l.get("qty", 0)
            price = l.get("price", 0)
            note  = l.get("note","").replace("\n"," | ")

            # Chiều cao dòng — đủ chứa ảnh 56px
            row_h(R, 46)

            vals = [stt, note, p.get("name",""), p.get("specs",""),
                    "", p.get("sku",""), p.get("supplier",""),
                    p.get("unit","Cái"), qty, price, ""]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(R, c, v)
                cell.font = font(False,9)
                cell.border = border()
                cell.alignment = align("left","center",True)
            # STT căn giữa
            ws.cell(R,1).alignment = align("center")
            ws.cell(R,9).alignment = align("center")
            ws.cell(R,10).number_format = '#,##0'
            ws.cell(R,11).number_format = '#,##0'
            ws.cell(R,11).value = f"=I{R}*J{R}"

            # Ảnh
            img_url = p.get("image","")
            xl_img = fetch_img(img_url, (54,54))
            if xl_img:
                col_e = get_column_letter(5)
                xl_img.anchor = f"{col_e}{R}"
                ws.add_image(xl_img)

            section_rows[-1][2].append(R)
            R+=1

        # Tổng giải pháp (công thức SUM)
        if section_rows[-1][2]:
            first, last = section_rows[-1][2][0], section_rows[-1][2][-1]
            ws.cell(sec_total_row, 11, f"=SUM(K{first}:K{last})")
            ws.cell(sec_total_row, 11).font = Font(name="Arial",bold=True,size=10,color=SEC_FG)
            ws.cell(sec_total_row, 11).number_format = '#,##0'
            ws.cell(sec_total_row, 11).alignment = align("right")

    R+=1

    # ══════════════════════════════════════════════════════════════════════════
    # BẢNG TỔNG HỢP CÁC GIẢI PHÁP
    # ══════════════════════════════════════════════════════════════════════════
    row_h(R,18)
    mr(R,1,11,"TỔNG HỢP CÁC GIẢI PHÁP NHÀ THÔNG MINH",
       font(True,12,WHITE), fill(BRAND), align("center"), border(BRAND))
    R+=1

    sum_rows = []
    for sec_name, sec_r, _ in section_rows:
        row_h(R,15)
        mr(R,1,10, sec_name, font(False,10), fill(GRAY_BG), align("left"), border())
        ws.cell(R,11, f"=K{sec_r}").number_format = "#,##0"
        ws.cell(R,11).font = font(True,10)
        ws.cell(R,11).alignment = align("right")
        ws.cell(R,11).border = border()
        ws.cell(R,11).fill = fill(GRAY_BG)
        sum_rows.append(R)
        R+=1

    # Tổng tiền hàng
    hang_r = R
    row_h(R,16)
    mr(R,1,10,"Tổng tiền hàng:", font(True,11), fill(BRAND_BG), align("left"), border(BRAND))
    sum_formula = "+".join([f"K{r}" for r in sum_rows])
    ws.cell(R,11, f"={sum_formula}").number_format = "#,##0"
    ws.cell(R,11).font = Font(name="Arial",bold=True,size=11,color=SEC_FG)
    ws.cell(R,11).alignment = align("right")
    ws.cell(R,11).fill = fill(BRAND_BG)
    ws.cell(R,11).border = border(BRAND)
    R+=1

    # Nhân công
    nc_pct = float(company.get("laborPercent",10)) / 100
    nc_r = R
    row_h(R,16)
    mr(R,1,10, f"Nhân công, lắp đặt & lập trình hệ thống ({company.get('laborPercent',10)}%):",
       font(True,11), fill(GRAY_BG), align("left"), border())
    ws.cell(R,11, f"=K{hang_r}*{nc_pct}").number_format = "#,##0"
    ws.cell(R,11).font = font(True,11)
    ws.cell(R,11).alignment = align("right")
    ws.cell(R,11).border = border()
    ws.cell(R,11).fill = fill(GRAY_BG)
    R+=1

    # Giá trị HĐ
    row_h(R,20)
    mr(R,1,10,"TỔNG GIÁ TRỊ HỢP ĐỒNG:", font(True,13,WHITE), fill(BRAND),
       align("left"), border(BRAND))
    ws.cell(R,11, f"=K{hang_r}+K{nc_r}").number_format = "#,##0"
    ws.cell(R,11).font = Font(name="Arial",bold=True,size=13,color=WHITE)
    ws.cell(R,11).alignment = align("right")
    ws.cell(R,11).fill = fill(BRAND)
    ws.cell(R,11).border = border(BRAND)
    R+=2

    # Ghi chú
    row_h(R,14)
    mr(R,1,11,"* Báo giá có giá trị trong vòng 14 ngày kể từ ngày báo giá.",
       font(False,9,TEXT2), fill(GRAY_BG), align("left"), border())
    R+=2

    # Ký tên
    row_h(R,14)
    mr(R,1,5,"KHÁCH HÀNG", font(True,10), None, align("center"))
    mr(R,6,11,"ĐẠI DIỆN CÔNG TY", font(True,10), None, align("center"))
    R+=1
    row_h(R,12)
    mr(R,1,5,"(Ký xác nhận & Ghi rõ họ tên)", font(False,8,TEXT2), None, align("center"))
    mr(R,6,11,company.get("name",""), font(False,8,TEXT2), None, align("center"))
    R+=4
    row_h(R,12)
    mr(R,1,5,"", None, None, align("center"))
    mr(R,6,11,f"(Ký và ghi rõ họ tên)", font(False,8,TEXT2), None, align("center"))

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A10"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


if __name__ == "__main__":
    # Test với data giả
    data = {
        "company": {
            "name": "CÔNG TY CỔ PHẦN NGUYÊN ĐÀ GROUP",
            "address": "119 Diệp Minh Châu, Hòa Xuân, Cẩm Lệ, Đà Nẵng",
            "taxCode": "0401940219",
            "phone": "0968 546 586",
            "website": "www.lumidanang.vn  |  www.nguyendagroup.com",
            "salesPerson": "Trần Trung Kiên",
            "salesPhone": "0968.546.586",
            "laborPercent": 10,
        },
        "customer": {
            "name": "CÔNG TRÌNH ANH PHÚC",
            "phone": "0901234567",
            "address": "Hội An",
            "quoteNumber": "BGAD-226-01/BG2026",
            "category": "Giải pháp nhà thông minh Lumi",
        },
        "rooms": [
            {
                "name": "I./ Giải pháp chiếu sáng tự động thông minh",
                "lines": [
                    {"productId":"p1","qty":19,"price":1026000,"note":"Tầng 1: 10 | Tầng 2: 7 | Tầng mái: 2",
                     "product":{"name":"Công tắc Lumes 1 nút","sku":"LM-1G2W-C(G)","supplier":"Lumi",
                                "unit":"Chiếc","specs":"2 dây · Zigbee · 100-240VAC",
                                "image":"https://lumi.vn/wp-content/uploads/2023/08/LM-1G2WC.png"}},
                    {"productId":"p2","qty":96,"price":378000,"note":"Tầng 1: 50 | Tầng 2: 38 | Tầng mái: 8",
                     "product":{"name":"Ổ cắm Lumes","sku":"LM-SK4/S-PC(G)","supplier":"Lumi",
                                "unit":"Chiếc","specs":"Vuông · champagne/darkgrey",
                                "image":""}},
                ],
            },
            {
                "name": "II./ Hệ thống cảm biến",
                "lines": [
                    {"productId":"p3","qty":24,"price":1620000,"note":"",
                     "product":{"name":"Cảm biến hiện diện BLE","sku":"LM-PCB-B","supplier":"Lumi",
                                "unit":"Bộ","specs":"Pin AAA · BLE Mesh","image":""}},
                ],
            },
        ],
        "calc": {"deviceTotal": 0, "laborTotal": 0, "grand": 0},
    }

    out = build_excel(data)
    with open("/mnt/user-data/outputs/test_baogia.xlsx","wb") as f:
        f.write(out)
    print(f"OK: {len(out)//1024}KB")
